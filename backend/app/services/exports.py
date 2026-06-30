from __future__ import annotations

import csv
import html
import json
import logging
import math
import re
import shutil
import uuid
from pathlib import Path
from typing import Any

import fitz

from backend.app.database import Database
from backend.app.models import ExportRecord, utc_now_iso
from backend.app.services.markup_memory import MarkupMemoryService
from backend.app.services.placement_coordinates import (
    image_rect_to_pdf_rect,
    pdf_rect_to_display_rect,
    round_rect,
)
from backend.app.services.review_coverage import project_review_coverage_summary
from backend.app.services.storage import require_project_source_pdf_path


logger = logging.getLogger(__name__)


class ExportService:
    def __init__(self, db: Database, data_dir: Path) -> None:
        self.db = db
        self.data_dir = Path(data_dir)

    def export_project(
        self,
        project_id: str,
        accepted_only: bool = True,
        statuses: list[str] | None = None,
        export_mode: str = "draft",
        reviewer_name: str | None = None,
        final_export_confirmed: bool = False,
        acknowledge_validation_warnings: bool = False,
    ) -> dict[str, Any]:
        project = self.db.get_project(project_id)
        source_pdf = require_project_source_pdf_path(self.data_dir, project_id, project.get("source_pdf_path"))

        mode = "final" if str(export_mode or "draft").strip().lower() == "final" else "draft"
        export_statuses = statuses if statuses is not None else (["accepted"] if accepted_only or mode == "final" else ["accepted", "needs_review"])
        if mode == "final":
            if not final_export_confirmed:
                self._record_export_blocked(project_id, "Final export requires reviewer confirmation.", mode)
                raise ValueError("Final export requires reviewer signoff confirmation.")
            if set(export_statuses) != {"accepted"}:
                self._record_export_blocked(project_id, "Final export may include accepted findings only.", mode)
                raise ValueError("Final export may include accepted findings only. Use draft export for other reviewer statuses.")
        findings = self.db.list_findings(project_id, statuses=export_statuses, sources=["ai"])
        sheets = self.db.list_sheets(project_id)
        review_coverage = project_review_coverage_summary(sheets, self.db.list_ai_import_batches(project_id, limit=1000))
        if mode == "final" and review_coverage["review_coverage_status"] != "complete":
            self._record_export_blocked(project_id, "Review coverage is incomplete.", mode, {"review_coverage": review_coverage})
            raise ValueError(
                "Final export blocked because review coverage is not complete. Import AI reviewed_pages confirmations for every page before final export."
            )
        sheet_by_id = {sheet["id"]: sheet for sheet in sheets}
        if not findings:
            status_label = ", ".join(export_statuses) if export_statuses else "selected"
            raise ValueError(
                f"No AI findings match the export status selection ({status_label}). Import or update findings before exporting."
            )
        if mode == "final":
            manual_status_count = sum(1 for finding in findings if finding.get("status") == "needs_manual_placement")
            existing_manual_placement_count = sum(1 for finding in findings if finding.get("placement_status") == "manual_placement_needed")
            if manual_status_count or existing_manual_placement_count:
                self._record_export_blocked(
                    project_id,
                    "Manual placement is still needed.",
                    mode,
                    {
                        "needs_manual_placement_status_count": manual_status_count,
                        "manual_placement_needed_count": existing_manual_placement_count,
                    },
                )
                raise ValueError("Final export blocked because one or more accepted findings still need manual placement.")

        export_id = str(uuid.uuid4())
        export_dir = self.data_dir / "projects" / project_id / "exports" / export_id
        export_dir.mkdir(parents=True, exist_ok=True)

        output_stem = f"{_safe_stem(project['name'])}_{mode}"
        marked_pdf = export_dir / f"{output_stem}_marked.pdf"
        csv_path = export_dir / f"{output_stem}_qc_log.csv"
        xlsx_path = export_dir / f"{output_stem}_qc_log.xlsx"
        json_path = export_dir / f"{output_stem}_findings.json"
        summary_path = export_dir / f"{output_stem}_review_summary.md"
        html_path = export_dir / f"{output_stem}_review_summary.html"
        signoff = {
            "reviewer_name": (reviewer_name or "Local reviewer").strip() or "Local reviewer",
            "timestamp": utc_now_iso(),
            "final_export_confirmed": bool(final_export_confirmed) if mode == "final" else False,
        }

        placement_results = self._write_marked_pdf(source_pdf, marked_pdf, findings)
        placement_summary = _placement_summary(placement_results)
        if mode == "final" and int(placement_summary.get("manual_placement_needed") or 0) > 0:
            self._record_export_blocked(project_id, "Generated PDF still has manual placement needed.", mode, {"placement_summary": placement_summary})
            _remove_export_dir(export_dir)
            raise ValueError("Final export blocked because generated placement validation found manual placement needed.")
        for finding in findings:
            placement = placement_results.get(finding.get("id")) or {}
            finding["placement_status"] = placement.get("placement_status")
            finding["placement_details"] = placement
            if finding.get("id") and placement.get("placement_status"):
                self.db.update_finding_placement(finding["id"], placement["placement_status"], placement)
        self._write_csv(csv_path, findings, sheet_by_id)
        self._write_xlsx(xlsx_path, findings, sheet_by_id)
        self._write_json(json_path, findings)
        self._write_summary(
            summary_path,
            html_path,
            project,
            sheets,
            findings,
            export_mode=mode,
            review_coverage=review_coverage,
            signoff=signoff,
            placement_summary=placement_summary,
            validation=None,
        )
        validation = validate_marked_pdf_export(
            source_pdf=source_pdf,
            marked_pdf=marked_pdf,
            expected_findings=len(findings),
            placement_summary=placement_summary,
        )
        if mode == "final" and validation["status"] == "failed":
            self._record_export_blocked(project_id, "Generated PDF validation failed.", mode, {"validation": validation})
            _remove_export_dir(export_dir)
            raise ValueError("Final export blocked because generated PDF validation failed.")
        if mode == "final" and validation["status"] == "warning" and not acknowledge_validation_warnings:
            self._record_export_blocked(project_id, "Generated PDF validation warnings were not acknowledged.", mode, {"validation": validation})
            _remove_export_dir(export_dir)
            raise ValueError("Final export blocked because generated PDF validation has warnings. Acknowledge validation warnings before final export.")
        self._write_summary(
            summary_path,
            html_path,
            project,
            sheets,
            findings,
            export_mode=mode,
            review_coverage=review_coverage,
            signoff=signoff,
            placement_summary=placement_summary,
            validation=validation,
        )

        record = ExportRecord(
            id=export_id,
            project_id=project_id,
            export_dir=str(export_dir),
            marked_pdf_path=str(marked_pdf),
            csv_path=str(csv_path),
            qa_report_path=str(csv_path),
            xlsx_path=str(xlsx_path) if xlsx_path.exists() else None,
            json_path=str(json_path),
            summary_path=str(summary_path),
            created_at=utc_now_iso(),
        ).model_dump()
        record["html_path"] = str(html_path)
        record["status_filter"] = export_statuses
        record["finding_count"] = len(findings)
        record["validation"] = validation
        record["export_mode"] = mode
        record["review_coverage"] = review_coverage
        record["signoff"] = signoff if mode == "final" else None
        self.db.insert_export(record)
        self.db.insert_project_event(
            project_id,
            "final_export_created" if mode == "final" else "draft_export_created",
            {
                "export_id": export_id,
                "status_filter": export_statuses,
                "finding_count": len(findings),
                "validation_status": validation.get("status"),
                "review_coverage_status": review_coverage.get("review_coverage_status"),
                "signoff": signoff if mode == "final" else None,
            },
        )
        try:
            MarkupMemoryService(self.db).collect_exported_findings(project_id, findings)
        except Exception as exc:
            logger.warning(
                "Markup Memory export capture failed for project_id=%s export_id=%s: %s",
                project_id,
                export_id,
                exc,
                exc_info=True,
            )
        return {
            "export": record,
            "findings_exported": len(findings),
            "placement_summary": placement_summary,
            "validation": validation,
            "export_mode": mode,
            "review_coverage": review_coverage,
            "signoff": signoff if mode == "final" else None,
        }

    def _record_export_blocked(self, project_id: str, reason: str, export_mode: str, details: dict[str, Any] | None = None) -> None:
        self.db.insert_project_event(
            project_id,
            "final_export_blocked" if export_mode == "final" else "draft_export_blocked",
            {"reason": reason, "export_mode": export_mode, **(details or {})},
        )

    def _write_marked_pdf(self, source_pdf: Path, target_pdf: Path, findings: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
        by_page: dict[int, list[dict[str, Any]]] = {}
        placement_results: dict[str, dict[str, Any]] = {}
        for finding in findings:
            page_number = _coerce_export_page_number(finding.get("page_number"))
            if page_number is None:
                placement_results[finding.get("id") or finding.get("stable_id") or str(len(placement_results))] = _placement_details(
                    "manual_placement_needed",
                    target_found=False,
                    exported=False,
                    note="Finding has no valid PDF page number.",
                )
                continue
            by_page.setdefault(page_number, []).append(finding)

        with fitz.open(source_pdf) as doc:
            for page_number, page_findings in by_page.items():
                if page_number < 1 or page_number > len(doc):
                    for finding in page_findings:
                        placement_results[finding.get("id") or finding.get("stable_id") or str(len(placement_results))] = _placement_details(
                            "manual_placement_needed",
                            target_found=False,
                            exported=False,
                            note=f"Page {page_number} is outside the source PDF page range.",
                        )
                    continue
                page = doc[page_number - 1]
                page_bounds = _page_coordinate_bounds(page)
                fallback_index = 0
                for finding in page_findings:
                    color = _severity_color(finding.get("severity"))
                    content = f"{finding.get('stable_id')}: {finding.get('comment_text')}"
                    placement = _finding_placement(page, finding)
                    rect = placement.get("rect")
                    if isinstance(rect, fitz.Rect):
                        cloud_created = _add_target_cloud(page, page_bounds, rect, color, finding)
                        placement["annotation_style"] = "cloud_plus_note" if cloud_created else "rectangle_note_fallback"
                        placement["target_cloud_created"] = cloud_created
                        point = _safe_note_point(page_bounds, rect.x1 + 8, rect.y0)
                    else:
                        placement["annotation_style"] = "sticky_note_fallback"
                        placement["target_cloud_created"] = False
                        content = f"{content}\nManual placement needed: target text was not found for automatic placement."
                        point = _fallback_note_point(page_bounds, fallback_index)
                        fallback_index += 1
                    note = page.add_text_annot(point, content)
                    note.set_info(
                        title="AutoQC",
                        subject=finding.get("category", "QC Finding"),
                        content=content,
                    )
                    note.set_colors(stroke=color)
                    note.update()
                    placement_results[finding.get("id") or finding.get("stable_id") or str(len(placement_results))] = {
                        key: value
                        for key, value in placement.items()
                        if key != "rect"
                    }
            doc.save(target_pdf, garbage=4, deflate=True)
        return placement_results

    def _write_csv(self, path: Path, findings: list[dict[str, Any]], sheet_by_id: dict[str, dict[str, Any]]) -> None:
        fieldnames = [
            "finding_id",
            "page_number",
            "drawing_number",
            "sheet_identifier",
            "category",
            "severity",
            "reviewer_status",
            "ai_source",
            "ai_batch_id",
            "prompt_version",
            "target_text",
            "evidence",
            "required_update",
            "rationale",
            "final_exported_comment",
            "comment",
            "placement_status",
            "target_text_found",
            "finding_exported",
            "manual_placement_needed",
            "reviewer_note",
            "confidence",
            "source",
            "created_at",
        ]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for finding in findings:
                sheet = sheet_by_id.get(finding.get("sheet_id") or "", {})
                writer.writerow(_qc_log_row(finding, sheet))

    def _write_xlsx(self, path: Path, findings: list[dict[str, Any]], sheet_by_id: dict[str, dict[str, Any]]) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception:
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "QC Log"
        headers = [
            "finding_id",
            "page_number",
            "drawing_number",
            "sheet_identifier",
            "category",
            "severity",
            "reviewer_status",
            "ai_source",
            "ai_batch_id",
            "prompt_version",
            "target_text",
            "evidence",
            "required_update",
            "rationale",
            "final_exported_comment",
            "comment",
            "placement_status",
            "target_text_found",
            "finding_exported",
            "manual_placement_needed",
            "reviewer_note",
            "confidence",
            "source",
            "created_at",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="253241")
        for finding in findings:
            row = _qc_log_row(finding, sheet_by_id.get(finding.get("sheet_id") or "", {}))
            sheet.append([row[key] for key in headers])
        widths = [16, 12, 18, 26, 24, 12, 16, 18, 22, 24, 46, 58, 42, 42, 46, 46, 22, 14, 14, 18, 32, 12, 12, 22]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        workbook.save(path)

    def _write_json(self, path: Path, findings: list[dict[str, Any]]) -> None:
        path.write_text(json.dumps(findings, indent=2, ensure_ascii=True), encoding="utf-8")

    def _write_summary(
        self,
        markdown_path: Path,
        html_path: Path,
        project: dict[str, Any],
        sheets: list[dict[str, Any]],
        findings: list[dict[str, Any]],
        *,
        export_mode: str,
        review_coverage: dict[str, Any],
        signoff: dict[str, Any],
        placement_summary: dict[str, int],
        validation: dict[str, Any] | None,
    ) -> None:
        severity_counts: dict[str, int] = {}
        category_counts: dict[str, int] = {}
        placement_counts = placement_summary or _placement_summary(
            {
                finding.get("id") or finding.get("stable_id") or str(index): finding.get("placement_details") or {}
                for index, finding in enumerate(findings)
            }
        )
        latest_import = next((batch for batch in self.db.list_ai_import_batches(project["id"]) if batch.get("import_status") == "imported"), None)
        latest_import_metadata = (latest_import.get("metadata") or {}) if latest_import else {}
        import_quality = latest_import_metadata.get("quality_report") if latest_import else None
        direct_review_warning = (
            "Direct AI Review findings are text-context-only and experimental; verify them against the attached/source PDF before relying on this export."
            if latest_import_metadata.get("direct_review_mode") == "text_context_only"
            else None
        )
        for finding in findings:
            severity_counts[finding["severity"]] = severity_counts.get(finding["severity"], 0) + 1
            category_counts[finding["category"]] = category_counts.get(finding["category"], 0) + 1

        lines = [
            f"# Review Summary: {project['name']}",
            "",
            f"- Project ID: `{project['id']}`",
            f"- Export mode: {export_mode.upper()}",
            f"- Sheets processed: {len(sheets)}",
            f"- Findings exported: {len(findings)}",
            f"- Severity counts: {severity_counts}",
            f"- Category counts: {category_counts}",
            f"- Placement summary: {placement_counts}",
            f"- Export validation: {(validation or {}).get('status', 'not run')}",
            f"- Export validation warnings: {(validation or {}).get('warnings', [])}",
            f"- Export validation errors: {(validation or {}).get('errors', [])}",
            f"- Review coverage: {review_coverage.get('review_coverage_status')} ({review_coverage.get('review_coverage_percent')}%)",
            f"- Expected review pages: {review_coverage.get('expected_review_pages')}",
            f"- Missing review pages: {review_coverage.get('missing_review_pages')}",
            f"- Reviewer signoff: {signoff.get('reviewer_name')} at {signoff.get('timestamp') if export_mode == 'final' else 'not required for draft'}",
            f"- Direct AI warning: {direct_review_warning or 'none'}",
            f"- Import quality summary: {import_quality or 'No imported AI quality report available'}",
            "",
            "## Exported Findings",
            "",
        ]
        for finding in findings:
            sheet = next((item for item in sheets if item["id"] == finding.get("sheet_id")), {})
            lines.extend(
                [
                    f"### {finding['stable_id']} - {finding['title']}",
                    "",
                    f"- Severity: {finding['severity']}",
                    f"- Status: {finding.get('status', 'unknown')}",
                    f"- Category: {finding['category']}",
                    f"- Sheet: {sheet.get('drawing_number', 'UNKNOWN')} page {finding.get('page_number')}",
                    f"- Target text: {_target_text(finding) or 'Not found/provided'}",
                    f"- Placement: {finding.get('placement_status') or 'unknown'}",
                    f"- Comment: {finding['comment_text']}",
                    f"- Suggested correction: {finding['suggested_correction']}",
                    f"- Evidence: {_evidence_text(finding)}",
                    "",
                ]
            )
        markdown = "\n".join(lines)
        markdown_path.write_text(markdown, encoding="utf-8")
        html_lines = [
            "<!doctype html><html><head><meta charset='utf-8'><title>AutoQC Review Summary</title>",
            "<style>body{font-family:Arial,sans-serif;max-width:980px;margin:32px auto;line-height:1.45;color:#1f2933}",
            "h1,h2,h3{margin-top:24px} code{background:#eef2f5;padding:2px 4px;border-radius:3px}</style></head><body>",
        ]
        for line in lines:
            if line.startswith("# "):
                html_lines.append(f"<h1>{html.escape(line[2:])}</h1>")
            elif line.startswith("## "):
                html_lines.append(f"<h2>{html.escape(line[3:])}</h2>")
            elif line.startswith("### "):
                html_lines.append(f"<h3>{html.escape(line[4:])}</h3>")
            elif line.startswith("- "):
                html_lines.append(f"<p>{html.escape(line)}</p>")
            elif line:
                html_lines.append(f"<p>{html.escape(line)}</p>")
        html_lines.append("</body></html>")
        html_path.write_text("\n".join(html_lines), encoding="utf-8")


def _qc_log_row(finding: dict[str, Any], sheet: dict[str, Any]) -> dict[str, Any]:
    placement = finding.get("placement_details") if isinstance(finding.get("placement_details"), dict) else {}
    evidence_items = _evidence_items(finding)
    first_evidence = evidence_items[0] if evidence_items else {}
    return {
        "finding_id": finding.get("stable_id"),
        "page_number": finding.get("page_number"),
        "drawing_number": sheet.get("drawing_number", "UNKNOWN"),
        "sheet_identifier": sheet.get("sheet_title", "Unknown Sheet"),
        "category": finding.get("category"),
        "severity": finding.get("severity"),
        "reviewer_status": finding.get("status"),
        "ai_source": first_evidence.get("source_type") or ("manual_chat_prompt" if finding.get("source") == "ai" else finding.get("source")),
        "ai_batch_id": finding.get("ai_batch_id") or first_evidence.get("ai_batch_id"),
        "prompt_version": finding.get("prompt_version") or first_evidence.get("prompt_version"),
        "target_text": _target_text(finding),
        "evidence": _evidence_text(finding),
        "required_update": finding.get("suggested_correction"),
        "rationale": finding.get("reasoning_summary"),
        "final_exported_comment": finding.get("comment_text"),
        "comment": finding.get("comment_text"),
        "placement_status": finding.get("placement_status") or placement.get("placement_status"),
        "target_text_found": placement.get("target_found"),
        "finding_exported": placement.get("exported"),
        "manual_placement_needed": placement.get("manual_placement_needed"),
        "reviewer_note": finding.get("reviewer_note"),
        "confidence": finding.get("confidence"),
        "source": finding.get("source"),
        "created_at": finding.get("created_at"),
    }


def validate_marked_pdf_export(
    *,
    source_pdf: Path,
    marked_pdf: Path,
    expected_findings: int,
    placement_summary: dict[str, int],
) -> dict[str, Any]:
    checks: list[dict[str, Any]] = []
    warnings: list[str] = []
    errors: list[str] = []
    annotation_count = 0
    source_page_count: int | None = None
    marked_page_count: int | None = None

    if not marked_pdf.exists():
        errors.append("Marked PDF file was not created.")
        return _validation_result("failed", checks, warnings, errors, expected_findings, annotation_count, source_page_count, marked_page_count, placement_summary)

    try:
        with fitz.open(source_pdf) as source_doc, fitz.open(marked_pdf) as marked_doc:
            source_page_count = len(source_doc)
            marked_page_count = len(marked_doc)
            checks.append({"name": "file_exists", "passed": True, "detail": str(marked_pdf)})
            checks.append(
                {
                    "name": "page_count_matches",
                    "passed": source_page_count == marked_page_count,
                    "detail": f"source={source_page_count}, marked={marked_page_count}",
                }
            )
            if source_page_count != marked_page_count:
                errors.append("Marked PDF page count does not match the source PDF.")
            for page in marked_doc:
                annotation_count += sum(1 for _ in (page.annots() or []))
    except Exception as exc:
        errors.append(f"Marked PDF could not be reopened for validation: {exc}")
        return _validation_result("failed", checks, warnings, errors, expected_findings, annotation_count, source_page_count, marked_page_count, placement_summary)

    checks.append(
        {
            "name": "annotations_present",
            "passed": annotation_count > 0,
            "detail": f"{annotation_count} annotations found",
        }
    )
    if expected_findings > 0 and annotation_count == 0:
        errors.append("No PDF annotations were found for the exported findings.")
    elif annotation_count < expected_findings:
        warnings.append(f"Only {annotation_count} annotations were found for {expected_findings} exported findings.")

    checks.append(
        {
            "name": "expected_finding_count",
            "passed": expected_findings > 0,
            "detail": f"{expected_findings} findings selected by export status filter",
        }
    )
    if expected_findings <= 0:
        errors.append("Export had zero selected findings.")

    manual = int(placement_summary.get("manual_placement_needed") or 0)
    page_level = int(placement_summary.get("page_level_fallback") or 0)
    if manual:
        warnings.append(f"{manual} exported findings still need manual placement.")
    if page_level:
        warnings.append(f"{page_level} exported findings used page-level fallback notes.")

    status = "failed" if errors else ("warning" if warnings else "passed")
    return _validation_result(status, checks, warnings, errors, expected_findings, annotation_count, source_page_count, marked_page_count, placement_summary)


def _validation_result(
    status: str,
    checks: list[dict[str, Any]],
    warnings: list[str],
    errors: list[str],
    expected_findings: int,
    annotation_count: int,
    source_page_count: int | None,
    marked_page_count: int | None,
    placement_summary: dict[str, int],
) -> dict[str, Any]:
    return {
        "status": status,
        "checks": checks,
        "warnings": warnings,
        "errors": errors,
        "expected_findings": expected_findings,
        "annotation_count": annotation_count,
        "source_page_count": source_page_count,
        "marked_page_count": marked_page_count,
        "placement_summary": placement_summary,
    }


def _target_text(finding: dict[str, Any]) -> str:
    for item in _evidence_items(finding):
        for key in ["target_text", "markup_text", "text_excerpt"]:
            value = item.get(key)
            if isinstance(value, str) and value.strip():
                return " ".join(value.split())
    return ""


def _evidence_text(finding: dict[str, Any]) -> str:
    parts: list[str] = []
    for item in _evidence_items(finding):
        for key in ["observation", "target_text", "text_excerpt", "required_update"]:
            value = item.get(key)
            if isinstance(value, str) and value.strip():
                clean = " ".join(value.split())
                if clean not in parts:
                    parts.append(clean)
    return " | ".join(parts)


def _evidence_items(finding: dict[str, Any]) -> list[dict[str, Any]]:
    return [item for item in finding.get("evidence", []) or [] if isinstance(item, dict)]


def _severity_color(severity: str | None) -> tuple[float, float, float]:
    return {
        "Critical": (0.85, 0.05, 0.05),
        "Major": (1.0, 0.45, 0.05),
        "Minor": (0.95, 0.78, 0.05),
        "Note": (0.1, 0.35, 0.8),
    }.get(severity or "", (0.2, 0.2, 0.2))


def _coerce_export_page_number(value: Any) -> int | None:
    try:
        page_number = int(value or 1)
    except (TypeError, ValueError):
        return None
    return page_number if page_number >= 1 else None


def _add_target_cloud(
    page: fitz.Page,
    page_bounds: fitz.Rect,
    target_rect: fitz.Rect,
    color: tuple[float, float, float],
    finding: dict[str, Any],
) -> bool:
    cloud_rect = _pad_rect(target_rect, page_bounds, 4)
    if cloud_rect is None:
        return False

    cloud = page.add_rect_annot(cloud_rect)
    cloud.set_info(
        title="AutoQC",
        subject=f"Target cloud - {finding.get('category', 'QC Finding')}",
        content=f"Clouded target for {finding.get('stable_id') or finding.get('id') or 'AutoQC finding'}",
    )
    cloud.set_colors(stroke=color)
    cloud.set_opacity(0.95)
    try:
        cloud.set_border(width=1.5, clouds=2)
    except TypeError:
        cloud.set_border(width=1.5)
    except Exception:
        pass
    cloud.update()
    return True


def _location_rect(location: Any) -> fitz.Rect | None:
    if not location:
        return None
    try:
        if isinstance(location, dict):
            if str(location.get("coordinate_space") or "") == "image_pixel":
                return None
            if all(key in location for key in ("x0", "y0", "x1", "y1")):
                return fitz.Rect(location["x0"], location["y0"], location["x1"], location["y1"])
            bbox = location.get("bbox") or location.get("rect")
            if isinstance(bbox, list) and len(bbox) >= 4:
                return fitz.Rect(bbox[0], bbox[1], bbox[2], bbox[3])
        if isinstance(location, list) and len(location) >= 4:
            return fitz.Rect(location[0], location[1], location[2], location[3])
    except Exception:
        return None
    return None


def _finding_placement(page: fitz.Page, finding: dict[str, Any]) -> dict[str, Any]:
    page_bounds = _page_coordinate_bounds(page)
    manual_image = _manual_image_placement(page, finding, page_bounds)
    if manual_image is not None:
        return manual_image

    location_rect = _safe_annotation_rect(_location_rect(finding.get("location")), page_bounds)
    if location_rect is not None:
        status = "manual_placement" if finding.get("placement_status") == "manual_placement" else "exact_target_found"
        return _placement_details(
            status,
            target_found=True,
            exported=True,
            rect=location_rect,
            page=page,
            note="Used reviewer manual placement." if status == "manual_placement" else "Used existing finding location.",
            method="manual_placement" if status == "manual_placement" else "provided_location",
        )

    search = _evidence_search_result(page, finding)
    if search.get("rect") is not None:
        return search

    has_target = bool(_search_candidates(finding))
    status = "page_level_fallback" if has_target else "manual_placement_needed"
    return _placement_details(
        status,
        target_found=False,
        exported=True,
        note="Target text was not found; added a page-level note." if has_target else "No target text was available; added a page-level note.",
        method="page_note",
    )


def _manual_image_placement(page: fitz.Page, finding: dict[str, Any], page_bounds: fitz.Rect) -> dict[str, Any] | None:
    placement = finding.get("placement_details")
    location = finding.get("location")
    image_rect = None
    image_width = None
    image_height = None

    if isinstance(placement, dict) and str(placement.get("coordinate_space") or "") == "image_pixel":
        image_rect = _coerce_rect_values(placement.get("manual_image_rect_json"))
        image_width = _coerce_positive_number(placement.get("image_width"))
        image_height = _coerce_positive_number(placement.get("image_height"))

    if image_rect is None and isinstance(location, dict) and str(location.get("coordinate_space") or "") == "image_pixel":
        image_rect = _coerce_rect_values(location.get("manual_image_rect") or location.get("bbox"))
        image_width = image_width or _coerce_positive_number(location.get("image_width"))
        image_height = image_height or _coerce_positive_number(location.get("image_height"))

    if image_rect is None or not image_width or not image_height:
        return None

    source_width = float(page.cropbox.width)
    source_height = float(page.cropbox.height)
    display_width = float(page.rect.width)
    display_height = float(page.rect.height)
    rotation = int(page.rotation or 0) % 360
    pdf_rect_json = image_rect_to_pdf_rect(
        image_rect,
        image_width=image_width,
        image_height=image_height,
        source_width=source_width,
        source_height=source_height,
        display_width=display_width,
        display_height=display_height,
        rotation=rotation,
    )
    pdf_rect = _safe_annotation_rect(fitz.Rect(pdf_rect_json), page_bounds)
    if pdf_rect is None:
        return None

    details = _placement_details(
        "manual_placement",
        target_found=True,
        exported=True,
        rect=pdf_rect,
        page=page,
        note="Used reviewer manual image-pixel placement.",
        method="manual_image_pixel_placement",
    )
    details["coordinate_space"] = "image_pixel"
    details["manual_image_rect_json"] = round_rect(image_rect)
    details["pdf_rect_json"] = details["rect_json"]
    details["display_rect_json"] = round_rect(image_rect)
    details["page_display_rect_json"] = pdf_rect_to_display_rect(
        details["rect_json"],
        source_width=source_width,
        source_height=source_height,
        rotation=rotation,
    )
    details["image_width"] = round(float(image_width), 2)
    details["image_height"] = round(float(image_height), 2)
    details["page_display_width"] = round(display_width, 2)
    details["page_display_height"] = round(display_height, 2)
    return details


def _evidence_search_rect(page: fitz.Page, finding: dict[str, Any]) -> fitz.Rect | None:
    result = _evidence_search_result(page, finding)
    rect = result.get("rect")
    return rect if isinstance(rect, fitz.Rect) else None


def _evidence_search_result(page: fitz.Page, finding: dict[str, Any]) -> dict[str, Any]:
    candidates = _search_candidates(finding)
    page_bounds = _page_coordinate_bounds(page)
    for phrase in candidates:
        try:
            matches = page.search_for(phrase)
        except Exception:
            matches = []
        for match in matches:
            rect = _pad_rect(match, page_bounds, 5)
            if rect is not None:
                return _placement_details(
                    "exact_target_found",
                    target_found=True,
                    exported=True,
                    rect=rect,
                    page=page,
                    matched_text=phrase,
                    method="exact_text_search",
                )

    for phrase in _fuzzy_search_candidates(candidates):
        try:
            matches = page.search_for(phrase)
        except Exception:
            matches = []
        for match in matches:
            rect = _pad_rect(match, page_bounds, 5)
            if rect is not None:
                return _placement_details(
                    "fuzzy_target_found",
                    target_found=True,
                    exported=True,
                    rect=rect,
                    page=page,
                    matched_text=phrase,
                    method="fuzzy_text_search",
                )

    return _placement_details("page_level_fallback", target_found=False, exported=True, method="page_note")


def _placement_summary(placement_results: dict[str, dict[str, Any]]) -> dict[str, int]:
    summary = {
        "exact_target_found": 0,
        "fuzzy_target_found": 0,
        "page_level_fallback": 0,
        "manual_placement": 0,
        "manual_placement_needed": 0,
        "target_cloud_created": 0,
        "cloud_plus_note": 0,
        "rectangle_note_fallback": 0,
        "sticky_note_fallback": 0,
    }
    for placement in placement_results.values():
        status = str(placement.get("placement_status") or "manual_placement_needed")
        summary[status] = int(summary.get(status, 0)) + 1
        if placement.get("target_cloud_created"):
            summary["target_cloud_created"] += 1
        annotation_style = str(placement.get("annotation_style") or "")
        if annotation_style:
            summary[annotation_style] = int(summary.get(annotation_style, 0)) + 1
    return summary


def _placement_details(
    placement_status: str,
    *,
    target_found: bool,
    exported: bool,
    rect: fitz.Rect | None = None,
    page: fitz.Page | None = None,
    matched_text: str | None = None,
    method: str | None = None,
    note: str | None = None,
) -> dict[str, Any]:
    details: dict[str, Any] = {
        "placement_status": placement_status,
        "target_found": target_found,
        "exported": exported,
        "manual_placement_needed": placement_status in {"page_level_fallback", "manual_placement_needed"},
    }
    if method:
        details["method"] = method
    if matched_text:
        details["matched_text"] = matched_text
    if note:
        details["note"] = note
    if rect is not None:
        details["rect"] = rect
        details["rect_json"] = [round(float(rect.x0), 2), round(float(rect.y0), 2), round(float(rect.x1), 2), round(float(rect.y1), 2)]
        if page is not None:
            details["coordinate_space"] = "pdf_unrotated"
            details["page_rotation"] = int(page.rotation or 0)
            details["source_width"] = round(float(page.cropbox.width), 2)
            details["source_height"] = round(float(page.cropbox.height), 2)
            details["display_width"] = round(float(page.rect.width), 2)
            details["display_height"] = round(float(page.rect.height), 2)
            details["display_rect_json"] = _rect_to_display_json(rect, page)
    return details


def _search_candidates(finding: dict[str, Any]) -> list[str]:
    candidates: list[str] = []
    for evidence in _evidence_items(finding):
        for key in ["target_text", "markup_text", "text_excerpt"]:
            phrase = _clean_search_text(evidence.get(key))
            if phrase and phrase not in candidates:
                candidates.append(phrase)
    return candidates


def _fuzzy_search_candidates(candidates: list[str]) -> list[str]:
    out: list[str] = []
    for candidate in candidates:
        for quoted in re.findall(r'"([^"]{2,80})"|' + r"'([^']{2,80})'", candidate):
            phrase = quoted[0] or quoted[1]
            if phrase and phrase not in out and phrase not in candidates:
                out.append(phrase)
        words = re.findall(r"[A-Za-z0-9\"'-]+", candidate)
        if len(words) >= 2:
            for count in [8, 6, 4, 3, 2]:
                if len(words) >= count:
                    phrase = " ".join(words[:count])
                    if phrase not in out and phrase not in candidates:
                        out.append(phrase)
                    break
        compact = re.sub(r"[^A-Za-z0-9]+", " ", candidate).strip()
        if compact and compact != candidate and compact not in out and compact not in candidates:
            out.append(compact)
    return out


def _search_phrase(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    clean = " ".join(value.split()).strip()
    if not clean:
        return None
    if len(clean) <= 80:
        return clean
    quoted = re.findall(r"'([^']{2,80})'", clean)
    if quoted:
        return quoted[0]
    words = clean.split()
    return " ".join(words[:8]) if len(words) >= 2 else None


def _clean_search_text(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    clean = " ".join(value.split()).strip()
    return clean or None


def _page_coordinate_bounds(page: fitz.Page) -> fitz.Rect:
    return fitz.Rect(0, 0, float(page.cropbox.width), float(page.cropbox.height))


def _rect_to_display_json(rect: fitz.Rect, page: fitz.Page) -> list[float]:
    return pdf_rect_to_display_rect(
        [float(rect.x0), float(rect.y0), float(rect.x1), float(rect.y1)],
        source_width=float(page.cropbox.width),
        source_height=float(page.cropbox.height),
        rotation=int(page.rotation or 0) % 360,
    )


def _coerce_rect_values(value: Any) -> list[float] | None:
    if not isinstance(value, list) or len(value) < 4:
        return None
    try:
        rect = [float(item) for item in value[:4]]
    except (TypeError, ValueError):
        return None
    if any(not math.isfinite(item) for item in rect):
        return None
    return round_rect(rect)


def _coerce_positive_number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) and number > 0 else None


def _pad_rect(rect: fitz.Rect, bounds: fitz.Rect, padding: float) -> fitz.Rect | None:
    padded = fitz.Rect(
        rect.x0 - padding,
        rect.y0 - padding,
        rect.x1 + padding,
        rect.y1 + padding,
    )
    return _safe_annotation_rect(padded, bounds)


def _fallback_note_point(bounds: fitz.Rect, index: int) -> fitz.Point:
    page_bounds = fitz.Rect(bounds)
    step = 32.0
    top = page_bounds.y0 + 48.0
    bottom = page_bounds.y1 - 24.0
    usable_height = max(step, bottom - top)
    rows_per_column = max(1, int(usable_height // step) + 1)
    row = index % rows_per_column
    column = index // rows_per_column
    return _safe_note_point(page_bounds, page_bounds.x1 - 36.0 - column * 36.0, top + row * step)


def _safe_note_point(bounds: fitz.Rect, x: float, y: float, margin: float = 24.0) -> fitz.Point:
    page_bounds = fitz.Rect(bounds)
    if page_bounds.is_empty or page_bounds.is_infinite:
        return fitz.Point(0, 0)

    min_x = page_bounds.x0 + margin
    max_x = page_bounds.x1 - margin
    min_y = page_bounds.y0 + margin
    max_y = page_bounds.y1 - margin
    if max_x < min_x:
        min_x = max_x = page_bounds.x0 + page_bounds.width / 2
    if max_y < min_y:
        min_y = max_y = page_bounds.y0 + page_bounds.height / 2

    safe_x = x if math.isfinite(float(x)) else max_x
    safe_y = y if math.isfinite(float(y)) else min_y
    return fitz.Point(min(max(safe_x, min_x), max_x), min(max(safe_y, min_y), max_y))


def _safe_annotation_rect(rect: Any, bounds: fitz.Rect, min_size: float = 4.0) -> fitz.Rect | None:
    if rect is None:
        return None
    try:
        candidate = fitz.Rect(rect)
        page_bounds = fitz.Rect(bounds)
    except Exception:
        return None
    values = (candidate.x0, candidate.y0, candidate.x1, candidate.y1, page_bounds.x0, page_bounds.y0, page_bounds.x1, page_bounds.y1)
    if any(not math.isfinite(float(value)) for value in values):
        return None
    if candidate.is_empty or candidate.is_infinite or page_bounds.is_empty or page_bounds.is_infinite:
        return None

    x0 = max(page_bounds.x0, min(candidate.x0, candidate.x1))
    y0 = max(page_bounds.y0, min(candidate.y0, candidate.y1))
    x1 = min(page_bounds.x1, max(candidate.x0, candidate.x1))
    y1 = min(page_bounds.y1, max(candidate.y0, candidate.y1))
    clipped = fitz.Rect(x0, y0, x1, y1)
    if clipped.is_empty or clipped.width <= 0 or clipped.height <= 0:
        return None

    if clipped.width < min_size or clipped.height < min_size:
        center = clipped.tl + (clipped.br - clipped.tl) * 0.5
        half_width = max(min_size / 2, clipped.width / 2)
        half_height = max(min_size / 2, clipped.height / 2)
        clipped = fitz.Rect(center.x - half_width, center.y - half_height, center.x + half_width, center.y + half_height)
        clipped = fitz.Rect(
            max(page_bounds.x0, clipped.x0),
            max(page_bounds.y0, clipped.y0),
            min(page_bounds.x1, clipped.x1),
            min(page_bounds.y1, clipped.y1),
        )
        if clipped.is_empty or clipped.width <= 0 or clipped.height <= 0:
            return None
    return clipped


def _safe_stem(name: str) -> str:
    value = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in name.strip())
    return value.strip("_") or "autoqc_review"


def _remove_export_dir(path: Path) -> None:
    try:
        if path.exists():
            shutil.rmtree(path)
    except OSError:
        logger.warning("Could not remove blocked export directory: %s", path, exc_info=True)
